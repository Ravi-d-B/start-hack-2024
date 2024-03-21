import streamlit as st
import pandas as pd
from streamlit_modal import Modal

from openpyxl.styles import Alignment, Font, PatternFill

from app.streamlit_app.database import *

st.set_page_config(
    layout="wide",
)

st.title("Test Grading")


class Test:
    def __init__(self, id, name, evaluations):
        self.id = id
        self.name = name
        self.evaluations = evaluations

    def get_id(self):
        return self.id

    def get_name(self):
        return self.name

    def get_evaluations(self):
        return self.evaluations


class StudentAnwsers:
    def __init__(self, student_id, anwsers):
        self.student_id = student_id
        self.anwsers = anwsers

    def get_id(self):
        return self.student_id

    def get_anwsers(self):
        return self.anwsers


loaded_tests = get_tests()
current_test = loaded_tests[0].id

tests_display = []


for test in loaded_tests:
    test_competencies = {}
    for competency in get_test_competencies(test.id):
        competency_id = competency.id
        competency_type = competency.get_competency_type().type
        test_competencies[competency_type] = competency_id

    tests_display.append(Test(test.id, test.test_name, test_competencies.keys()))

test_evaluations = {test.get_name(): test.get_evaluations() for test in tests_display}


# Use the select box widget
option = st.selectbox(
    'Select Test for Grading',
    test_evaluations.keys())

for test in tests_display:
    if test.get_name() == option:
        current_test = test.id


students = get_test_students(current_test)

# Create Student Marks Array
all_student_results = []

# Create a list of dictionaries, with each dictionary representing a row
data = []

for i, student in enumerate(students):
    data.clear()
    st.subheader(f"{student.name} - {student.id} ")
    for (question) in zip(test_evaluations[option]):
        row = {
            "Evaluations": question,
            "Das Klappt noch nicht": False,
            "Das Gelingt mit teilweise": False,
            "Das kann ich gut": False,
            "Das kann ich sehr gut": False}
        data.append(row)
    df = pd.DataFrame(data)
    question_answers = st.data_editor(df, use_container_width=True, disabled="col1", hide_index="true", height=None, key=i)
    all_student_results.append(StudentAnwsers(student.id, question_answers))


def get_mark(result):
    if result.iloc[1]:
        return 1
    if result.iloc[2]:
        return 2
    if result.iloc[3]:
        return 3
    if result.iloc[4]:
        return 4
    else:
        return 0

def get_competency_id(result):
    result_type = result.iloc[0][0]
    return test_competencies[result_type]

def saveData():
    for student_temp in all_student_results:
        for index in range(len(student_temp.get_anwsers())):
            question = student_temp.get_anwsers().iloc[index]
            add_test_evaluation_to_student(student_temp.get_id(), get_competency_id(question), get_mark(question), "")
    # modal.close()


if(st.button("Save")):
    saveData()

def convert_boolean(value):
    return "X" if value else ""

def prepare_dataframe(df):
     # Adjust 'Evaluations' to take the first element of the tuple
    df['Evaluations'] = df['Evaluations'].apply(lambda x: x[0])
    # Select columns to convert, assuming the first column 'Evaluations' should not be converted
    columns_to_convert = df.columns[1:]
    df[columns_to_convert] = df[columns_to_convert].applymap(convert_boolean)
    return df

def export_to_excel(students, student_marks):
   excel_file_path = "student_evaluations.xlsx"
   with pd.ExcelWriter(excel_file_path, engine='openpyxl') as writer:
    for student, df in zip(students, student_marks):
        print('df_answers', df.get_anwsers())
        prepared_df = prepare_dataframe(df.get_anwsers().copy())
        print(prepared_df)
        sheet_name = student.name[:31]

        # Write DataFrame to an Excel sheet
        prepared_df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Get the workbook and the worksheet objects
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]


        bold_font = Font(bold=True)
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        index = 0
        for col in worksheet.columns:
            if index == 0:
                pass
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                # Set center alignment for all cells
                cell.alignment = center_alignment
                
                # Apply bold font only to cells with 'X'
                if cell.value == "X":
                    cell.font = bold_font

                # Calculate max length for auto-adjustment
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column].width = adjusted_width

        LEVEL_1_COLOUR = 'FFF4B484'
        LEVEL_2_COLOUR = 'FFA8C98C'
        LEVEL_3_COLOUR = 'FFFA640C'
        LEVEL_4_COLOUR = 'FFFADA63'

        # Create fill objects for each level
        level_1_fill = PatternFill(start_color=LEVEL_1_COLOUR, end_color=LEVEL_1_COLOUR, fill_type="solid")
        level_2_fill = PatternFill(start_color=LEVEL_2_COLOUR, end_color=LEVEL_2_COLOUR, fill_type="solid")
        level_3_fill = PatternFill(start_color=LEVEL_3_COLOUR, end_color=LEVEL_3_COLOUR, fill_type="solid")
        level_4_fill = PatternFill(start_color=LEVEL_4_COLOUR, end_color=LEVEL_4_COLOUR, fill_type="solid")

        # Apply the fills to the first cell in columns B through E
        worksheet['B1'].fill = level_1_fill
        worksheet['C1'].fill = level_2_fill
        worksheet['D1'].fill = level_3_fill
        worksheet['E1'].fill = level_4_fill
        worksheet.column_dimensions['A'].width = 40



# Assuming 'students' is your list of student objects and 'student_marks' contains the edited DataFrames

if st.button("Export to Excel"):
    export_to_excel(students, all_student_results)  # This creates and prepares the Excel file

    with open("student_evaluations.xlsx", "rb") as file:
        st.download_button(
            label="Download Excel",
            data=file,
            file_name="student_evaluations.xlsx",
            mime="application/vnd.ms-excel"
        )

# if st.button('Confirm'):
#     saveData()

# modal = Modal(
#     "Do you want to confirm changes?",
#     key="confirm modal",
# )
#
# open_modal = st.button("Save")
# if open_modal:
#     modal.open()
#
#
# if modal.is_open():
#     with modal.container():
#         col1, col2, col3 = st.columns([6, 1, 1])
#         with col2:
#             if st.button('Confirm'):
#                 saveData()
#         with col3:
#             if st.button('Cancel'):
#                 modal.close()
