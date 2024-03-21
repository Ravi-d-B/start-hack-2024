# Function to display scores as a line plot
import streamlit as st
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import toml
from app.streamlit_app.database import get_student_tests
from app.streamlit_app.utils import get_prompt_template, client
from openpyxl.styles import Alignment, Font, PatternFill

from app.streamlit_app.database import (
    get_students, get_student_tests,
    get_student_test_evaluations, get_all_student_test_evaluations
)

from app.data.competencies import (
    get_all_subjects, get_compentencies_for_subject_code,
    get_full_comp_df
)

from app.streamlit_app.graph import create_graph

plt.rcParams['font.size'] = 24

# set font to Arial
plt.rcParams['font.sans-serif'] = 'Arial'

# Set the font globally
plt.rcParams['font.family'] = 'sans-serif'
plt.rcParams['font.sans-serif'] = ['Arial']  # Or any font you like

streamlit_theme = toml.load(".streamlit/config.toml")['theme']

st.title('Student Progression')


def plot_results(df):
    if df.empty:
        st.write("No data to display")
        return
    # Count number of unique concatenated categories
    cats = (df['cat1'] + df['cat2']).unique()
    num_cats = len(cats)

    # Calculate rows for n x 2 layout, rounding up. Ensure at least 1 row.
    num_rows = max(np.ceil(num_cats / 2).astype(int), 1)

    # Create subplots in a n x 2 layout, adjusting for when num_cats is 1 or 2
    plot_settings = {'sharex': False, 'sharey': True, 'dpi': 1200,
                     'facecolor': streamlit_theme['backgroundColor']}

    if num_cats <= 2:
        fig, axs = plt.subplots(num_rows, num_cats, figsize=(10*num_cats, 5*num_rows),**plot_settings)
        # Ensure axs is 2-dimensional
        axs = np.atleast_2d(axs)
    else:
        fig, axs = plt.subplots(num_rows, 2,figsize=(30, 7*num_rows), **plot_settings)

    # Flatten the axs array to simplify indexing
    axs = axs.flatten()

    for i, cat_val in enumerate(cats):
        # Filter on concatenated categories
        df_cat = df.query('(cat1 + cat2) == @cat_val')

        title_part_1 = df_cat['code'].iloc[0].rsplit('.', 2)[0:2]
        title_part_2 = df_cat['code'].iloc[0].rsplit('.', 1)[0]
        # Assuming get_full_comp_df() function fetches a DataFrame with 'code' and 'bezeichnung'
        # columns
        title_part_1 = get_full_comp_df().query('code == @title_part_1')['bezeichnung'].iloc[0]
        title_part_2 = get_full_comp_df().query('code == @title_part_2')['bezeichnung'].iloc[0]

        # Set title with the parts from the code
        axs[i].set_title(f'{title_part_1} - {title_part_2}', pad=20)

        # Make a line for every cat3
        for cat3 in df_cat['cat3'].unique():
            df_cat3 = df_cat.query('cat3 == @cat3')
            axs[i].plot(df_cat3['test_date'], df_cat3['score'],
                        marker='o',
                        label=df_cat3['code'].iloc[0])
        axs[i].legend()
        axs[i].set_facecolor('#f0f0f0')
        for side in ['top', 'bottom', 'left', 'right']:
            axs[i].spines[side].set(lw=2)

        # make line thicker
        for line in axs[i].lines:
            line.set_linewidth(6)
            line.set_markersize(13)
        axs[i].set_yticks([0, 1, 2, 3, 4, 5])
        axs[i].set_yticklabels(['',
                                'Das Klappt noch nicht',
                                'Das Gelingt mit teilweise',
                                'Das kann ich gut',
                                'Das kann ich sehr gut',
                                ''])

    # If num_cats is odd and greater than 2, hide the last subplot
    if num_cats % 2 != 0 and num_cats > 2:
        fig.delaxes(axs[-1])

    plt.tight_layout()
    st.pyplot(fig)


students = get_students()

# Dropdown of students
student = st.selectbox('Select a student',
                       students,
                       format_func=lambda student: student.name)
subjects = get_all_subjects()
if 'subject' not in st.session_state:
    st.session_state.subject = subjects[0]
st.session_state.subject = st.selectbox("Select a subject", subjects,
                                        index=subjects.index(st.session_state.subject))

evals = student.get_student_graph_data()

# Filter on subject
evals = evals.query('code.str.contains(@st.session_state.subject)')
# average scores for the same code on the same date
evals = evals.groupby(['test_date', 'code'])["score"].mean().reset_index()

# Split code into columns
evals["subjects"] = evals["code"].str.split(".").str[0]
evals["cat1"] = evals["code"].str.split(".").str[1]
evals["cat2"] = evals["code"].str.split(".").str[2]
evals["cat3"] = evals["code"].str.split(".").str[3]

plot_results(evals)

level_options = {
    " ": None,
    "Year 1": [2] + [None] * 6,
    "Year 2": [4, 4, 4, 1, 4, None, None],
    "Year 3": [4, 4, 4, 1, 4, 4, 3]
}

# Add a dropdown menu for selecting the level configuration
selected_option = st.selectbox(
    'Select the year for which you want to see the knowledge graph:',
    options=list(level_options.keys()),
    index=0
)

if selected_option != " ":
    # Get the selected levels configuration
    selected_levels = level_options[selected_option]

    # Create the graph based on the selected levels configuration
    graph = create_graph(selected_levels)

    # Display the graph
    st.graphviz_chart(graph, use_container_width=True)

information = student.get_student_graph_data()

if id not in st.session_state or id is not st.session_state.id:
    st.session_state[id] = get_prompt_template(student.name, information)
prompt_template = st.session_state[id]

# Display chat messages from history on app rerun
for message in prompt_template:
    if message["role"] != 'system':
        with st.chat_message(message["role"]):
            st.markdown(message["content"])

client = client()

if st.button("Ask AI for a student progress summary"):

    # Display assistant response in chat message container
    with st.chat_message("assistant"):
        stream = client.chat.completions.create(
            model="gpt-4-turbo-preview",
            messages=[
                {"role": m["role"], "content": m["content"]}
                for m in prompt_template
            ],
            temperature=0.2,
            stream=True
        )
        response = st.write_stream(stream)
    prompt_template.append({"role": "assistant", "content": response})

def convert_boolean(value):
    return "X" if value else ""

def prepare_dataframe(df):
     # Adjust 'Evaluations' to take the first element of the tuple
    df['Evaluations'] = df['Evaluations'].apply(lambda x: x[0])
    # Select columns to convert, assuming the first column 'Evaluations' should not be converted
    columns_to_convert = df.columns[1:]
    df[columns_to_convert] = df[columns_to_convert].applymap(convert_boolean)
    return df

def export_to_excel(student):
    excel_file_path = "student_evaluations.xlsx"
    with pd.ExcelWriter(excel_file_path, engine='openpyxl') as writer:
        tests = get_student_tests(student.id)
        for test in tests:
            # Retrieve evaluations for this test
            evaluations = get_student_test_evaluations(student.id, test.id)
            if(len(evaluations) == 0):
                pass

                        # Initialize a list to hold processed evaluation data
            processed_evaluations = []
            
            for eval in evaluations:
                # Assuming eval['score'] exists and is an integer 1-4
                score = eval.score
                competency = eval.get_test_competency()
                
                # Map score 1-4 to columns B-E with "X"
                # Create a row with 'None' for columns B-E and then replace the correct column with "X"
                row = [competency.get_competency_type().type] + [None, None, None, None]
                row[score] = "X"  # Place "X" in the column corresponding to the score
                
                processed_evaluations.append(row)
            
            # Convert the processed data into a DataFrame
            columns = ['Evaluation', 'Das Klappt noch nicht', 'Das Gelingt mit teilweise', 'Das kann ich gut', 'Das kann ich sehr gut']
            df = pd.DataFrame(processed_evaluations, columns=columns)
            # Generate a unique sheet name for this test, considering Excel's 31-character limit
            sheet_name = f"Test_{test.test_name}"[:31]
            # Write the DataFrame to a sheet in the Excel file
            df.to_excel(writer, sheet_name=sheet_name, index=False)

            worksheet = writer.sheets[sheet_name]
            bold_font = Font(bold=True)
            center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            index = 0
            for col in worksheet.columns:
                if index == 0:
                    index = index + 1
                    pass
                max_length = 0
                column = col[0].column_letter  # Get the column name
                for cell in col:
                    # Set center alignment for all cells
                    cell.alignment = center_alignment
                    
                    # Apply bold font only to cells with 'X'
                    if cell.value == "X":
                        cell.font = bold_font

                    # Calculate max length for auto-adjustment
                    try:  # Necessary to avoid error on empty cells
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[column].width = adjusted_width

            LEVEL_1_COLOUR = 'FFF4B484'
            LEVEL_2_COLOUR = 'FFA8C98C'
            LEVEL_3_COLOUR = 'FFFA640C'
            LEVEL_4_COLOUR = 'FFFADA63'

            # Create fill objects for each level
            level_1_fill = PatternFill(start_color=LEVEL_1_COLOUR, end_color=LEVEL_1_COLOUR, fill_type="solid")
            level_2_fill = PatternFill(start_color=LEVEL_2_COLOUR, end_color=LEVEL_2_COLOUR, fill_type="solid")
            level_3_fill = PatternFill(start_color=LEVEL_3_COLOUR, end_color=LEVEL_3_COLOUR, fill_type="solid")
            level_4_fill = PatternFill(start_color=LEVEL_4_COLOUR, end_color=LEVEL_4_COLOUR, fill_type="solid")

            # Apply the fills to the first cell in columns B through E
            worksheet['B1'].fill = level_1_fill
            worksheet['C1'].fill = level_2_fill
            worksheet['D1'].fill = level_3_fill
            worksheet['E1'].fill = level_4_fill
            worksheet.column_dimensions['A'].width = 40


if st.button("Export to Excel"):
    export_to_excel(student)  # This creates and prepares the Excel file

    with open("student_evaluations.xlsx", "rb") as file:
        st.download_button(
            label="Download Excel",
            data=file,
            file_name="student_evaluations.xlsx",
            mime="application/vnd.ms-excel"
        )
